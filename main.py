from ortools.constraint_solver import routing_enums_pb2
from ortools.constraint_solver import pywrapcp
import re
import numpy as np
import pandas as pd
import urllib
import matrix
import api
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from tabulate import tabulate
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def create_test_data_model():
    """Stores the data for the problem."""
    data = {}
    data["distance_matrix"] = [
        [   0, 100, 1000, 100, 1000],
        [ 100,   0,   50,  20,   50],
        [1000,  50,    0,  50,    0],
        [ 100,  20,   50,   0,   50],
        [1000,  50,    0,  50,    0],
    ]
    data["time_matrix"] = [
        [0, 10, 30, 10, 30],
        [10, 0,  5,  2,  5],
        [30, 5,  0,  5,  0],
        [10, 2,  5,  0,  5],
        [30, 5,  0,  5,  0],
    ]
    data["pickups_deliveries"] = [
        [1, 2],
        [3, 4],
    ]
    data["time_windows"] = [
        (0, 1000),
        (0, 1000),
        (470, 480),
        (0, 1000),
        (470, 480),
    ]
    data["num_vehicles"] = 2
    data["vehicle_capacities"] = [2, 2]
    data["demands"] = [0, 1, 0, 1, 0] # no of people at each node
    data["depot"] = 0
    return data


def create_data_model(drivers_df, workers_df, vehicle_capacity, timezone=None):
    '''Stores the data for the problem.'''
    data = {}

    start_of_day = time_to_epoch_time('00:00:00', timezone)
    end_of_day = time_to_epoch_time('23:59:59', timezone)
    # print(str(start_of_day))
    # print(str(end_of_day))
    data['addresses'] = []
    data['address_names'] = dict()
    address_dict = {} # maps index in data['addresses'] to full name of address
    address_rpt = [] # addresses with home and destinations repeated for the solver
    data['num_vehicles'] = drivers_df.shape[0]
    data['vehicle_capacities'] = [vehicle_capacity] * data['num_vehicles']
    data['depot'] = 0
    data['driver_names'] = drivers_df['Name'].tolist()
    data['worker_name'] = ['depot']
    data['node_type'] = ['depot']
    data['time_windows'] = [(start_of_day, end_of_day)]
    data['demands'] = [0]
    data['pickups_deliveries'] = []
    # data['time_matrix'] = []
    index = 1
    for _, row in workers_df.iterrows():
        # Home
        data['worker_name'].append(row['Name'])
        data['node_type'].append('from_1')
        home_address = format_address(row['Home'])
        data['demands'].append(1)
        data['time_windows'].append((start_of_day, end_of_day))
        row['Home Coordinates (Optional)'] = str(row['Home Coordinates (Optional)'])
        if row['Home Coordinates (Optional)'] != 'nan': # check if the coordinates cell isn't empty
            home_coor = format_address(row['Home Coordinates (Optional)'])
            data['address_names'][home_coor] = home_address
            home_address = home_coor
        if home_address not in address_dict:
            data['addresses'].append(home_address)
            address_dict[home_address] = len(data['addresses']) - 1

        # Work
        data['worker_name'].append(row['Name'])
        data['node_type'].append('to_1')
        dest_address = format_address(row['Destination'])
        dest_time = row['Arrival Time']
        # dest_time_sec = dest_time.hour * 3600 + dest_time.minute * 60 # + dest_time.second
        dest_time_sec = time_to_epoch_time(f'{dest_time.hour}:{dest_time.minute}:00', timezone)
        data['time_windows'].append((dest_time_sec-300, dest_time_sec))
        data['demands'].append(-1)
        row['Destination Coordinates (Optional)'] = str(row['Destination Coordinates (Optional)'])
        if row['Destination Coordinates (Optional)'] != 'nan': # check if the coordinates cell isn't empty
            dest_coor = format_address(row['Destination Coordinates (Optional)'])
            data['address_names'][dest_coor] = dest_address
            dest_address = dest_coor
        if dest_address not in address_dict:
            data['addresses'].append(dest_address)
            address_dict[dest_address] = len(data['addresses']) - 1

        address_rpt.append([dest_time.hour - 1, address_dict[home_address]])
        address_rpt.append([dest_time.hour, address_dict[dest_address]])

        data["pickups_deliveries"].append([index, index + 1])

        """Repeat, but in reverse for the return trip"""
        # Work
        data['worker_name'].append(row['Name'])
        data['node_type'].append('from_2')
        data['demands'].append(1)
        leave_time = row['Leave Time']
        # leave_time_sec = leave_time.hour * 3600 + leave_time.minute * 60 # + leave_time.second
        leave_time_sec = time_to_epoch_time(f'{leave_time.hour}:{leave_time.minute}:00', timezone)
        data['time_windows'].append((leave_time_sec, leave_time_sec + 300))

        # Home
        data['worker_name'].append(row['Name'])
        data['node_type'].append('to_2')
        dest_time = row['Arrival Time']
        data['time_windows'].append((start_of_day, end_of_day))
        data['demands'].append(-1)

        address_rpt.append([leave_time.hour, address_dict[dest_address]])
        address_rpt.append([leave_time.hour + 1, address_dict[home_address]])

        data["pickups_deliveries"].append([index + 2, index + 3])
        index += 4 

    data['address_rpt'] = [-1] + address_rpt
    data['address_dict'] = address_dict

    # Build time matrix from unique addresses
    data['API_key'] = api.get_api_key()
    # time_matrix = matrix.create_time_matrix(data, traffic = True, departure_time=str(time_to_epoch_time('7:30:00')))

    # Build time matrix for every hour of the next day
    time_matrices = []
    for hour in range(24):
        hour_str = f'{str(hour)}:00:00'
        time_matrices.append(matrix.create_time_matrix(data, traffic = True, departure_time=str(time_to_epoch_time(hour_str))))

    print()
    for index, tm in enumerate(time_matrices):
        print(index)
        print(tm)
        print()

    # Assemble time matrix for the solver

    array = np.zeros((len(address_rpt) + 1, len(address_rpt) + 1), dtype = int)

    for i in range(array.shape[0]):
        for j in range(array.shape[1]):
            if i!=0 and j!=0:
                array[i, j] = time_matrices[address_rpt[i-1][0]][address_rpt[i-1][1]][address_rpt[j-1][1]]

    data['time_matrix'] = array
    
    # print(address_dict)
    print(address_rpt)

    return data


def format_address(address):
    # Remove surrounding spaces
    address = address.strip()

    # Check if the address is in the format "double, double" or "double,double"
    pattern = r"^\d+\.\d+,\s?\d+\.\d+$"
    
    if re.match(pattern, address):
        # Remove the space if it exists
        address = re.sub(r"\s+", "", address)
    else:
        # Add ", Kuwait" to the end
        address = f"{address}, Kuwait"

    # Percent encode to use in URL
    return urllib.parse.quote(address)


def time_to_epoch_time(time_str, timezone=None):
    """
        Convert time from hh:mm:ss or hh:mm:ss am/pm to the UNIX timestamp of that time in the next day
    """
    time_str = time_str.upper().strip()
    split_time = time_str.split()
    time = split_time[0]
    time_hour, time_min, time_sec = time.split(':')
    time_hour = int(time_hour)
    time_min = int(time_min)
    time_sec = int(time_sec)
    if len(split_time) > 1:
        am_pm = split_time[1]
        if am_pm == 'AM' and time_hour == 12:
            time_hour = 0
        elif am_pm == 'PM' and time_hour > 12:
            time_hour += 12

    # print(time_str, time, time_hour, time_min)

    now = datetime.now() + timedelta(days=1)
    # print(now)

    if timezone:
        localized_now = now.astimezone(ZoneInfo(timezone))
    else:
        localized_now = now.astimezone()
    # print(localized_now)

    target_time = localized_now.replace(hour=time_hour, minute=time_min, second=time_sec, microsecond=0)
    target_timestamp = int(target_time.timestamp())
    # print(target_time)
    # print(target_timestamp)
    return target_timestamp


def epoch_time_to_time(timestamp, timezone=None):
    if timezone:
        dt = datetime.fromtimestamp(timestamp, ZoneInfo(timezone))
    else:
        dt = datetime.fromtimestamp(timestamp)
    # hour = dt.hour
    # minute = dt.minute
    # if hour == 0:
    #     hour = 12
    # elif hour > 12:
    #     hour -= 12
    # period = 'AM' if hour < 12 else 'PM'
    # time_str = f"{hour}:{minute} {period}"
    # return time_str
    # return dt.strftime('%m/%d/%Y  %I:%M %p')
    return dt.strftime('%I:%M %p')


def time_to_seconds(time_str):
    # Extract the AM/PM part
    period = time_str[-2:].upper()
    # Extract the hour and minute part
    time_part = time_str[:-2].strip()
    
    # Split the time into hours and minutes
    hours, minutes = map(int, time_part.split(':'))
    
    # Convert to 24-hour format if needed
    if period == 'PM' and hours != 12:
        hours += 12
    elif period == 'AM' and hours == 12:
        hours = 0

    # Calculate total seconds
    total_seconds = hours * 3600 + minutes * 60
    
    return total_seconds


def convert_seconds_to_hhmmss(seconds):
    # Calculate hours and minutes
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    seconds = seconds % 60
    # Determine AM or PM
    # period = 'AM' if hours < 12 else 'PM'
    # Adjust hours to 12-hour format
    # if hours == 0:
    #     hours = 12
    # elif hours > 12:
    #     hours -= 12

    return f"{hours:02}:{minutes:02}:{seconds:02}"


def convert_seconds_to_hhmm(seconds):
    # Calculate hours and minutes
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    # Determine AM or PM
    period = 'AM' if hours < 12 else 'PM'
    # Adjust hours to 12-hour format
    if hours == 0:
        hours = 12
    elif hours > 12:
        hours -= 12

    # Format the time string
    time_str = f"{hours:02}:{minutes:02} {period}"

    return time_str


def print_2d_matrix(matrix):
    # Convert each element to HH:MM:SS format
    formatted_matrix = [[convert_seconds_to_hhmmss(cell) for cell in row] for row in matrix]
    
    # Print the formatted matrix using tabulate
    print(tabulate(formatted_matrix, tablefmt="grid"))
    print()


def print_solution(data, manager, routing, solution, timezone):
    """Prints solution on console."""
    print('Time:')
    print(f"Objective: {solution.ObjectiveValue()}")
    time_dimension = routing.GetDimensionOrDie("Time")
    total_time = 0
    for vehicle_id in range(data["num_vehicles"]):
        index = routing.Start(vehicle_id)
        plan_output = f"Route for vehicle {vehicle_id}:\n"
        while not routing.IsEnd(index):
            time_var = time_dimension.CumulVar(index)
            plan_output += (
                f"{manager.IndexToNode(index)}"
                f" Time({epoch_time_to_time(solution.Min(time_var), timezone)},{epoch_time_to_time(solution.Max(time_var), timezone)})"
                # f" Time({solution.Min(time_var)},{solution.Max(time_var)})"
                " \n-> "
            )
            index = solution.Value(routing.NextVar(index))
        time_var = time_dimension.CumulVar(index)
        plan_output += (
            f"{manager.IndexToNode(index)}"
            f" Time({epoch_time_to_time(solution.Min(time_var), timezone)},{epoch_time_to_time(solution.Max(time_var), timezone)})\n"
            # f" Time({solution.Min(time_var)},{solution.Max(time_var)})\n"
        )
        plan_output += f"Time of the route: {epoch_time_to_time(solution.Min(time_var), timezone)}\n"
        print(plan_output)
        total_time += solution.Min(time_var)
    print(f"Total time of all routes: {total_time}")


def solve(data, same_driver=False, output_file='Schedule.xlsx', timezone=None):
    # Create the routing index manager.
    manager = pywrapcp.RoutingIndexManager(
        len(data["time_matrix"]), data["num_vehicles"], data["depot"]
    )

    # Create Routing Model.
    routing = pywrapcp.RoutingModel(manager)

    # Create and register a transit callback.
    # Define the cost of travel, to be the time taken to travel the arcs
    def time_callback(from_index, to_index):
        """Returns the travel time between the two nodes."""
        # Convert from routing variable Index to time matrix NodeIndex.
        from_node = manager.IndexToNode(from_index)
        to_node = manager.IndexToNode(to_index)
        return data["time_matrix"][from_node][to_node]

    transit_callback_index = routing.RegisterTransitCallback(time_callback)

    # Define cost of each arc to be the times taken to travel the arc
    routing.SetArcCostEvaluatorOfAllVehicles(transit_callback_index)

    # Add Time constraint.
    time = "Time"
    routing.AddDimension(
        transit_callback_index,
        18000000,  # allow waiting time
        10000000000,  # maximum time per vehicle
        False,  # Don't force start cumul to zero.
        time,
    )
    time_dimension = routing.GetDimensionOrDie(time)
    # set a large coefficient for the global span of the routes (i,e, maximum time of routes)
    # to minimize time of longest route
    time_dimension.SetGlobalSpanCostCoefficient(100)

    # Add time window constraints for each location except depot.
    for location_idx, time_window in enumerate(data["time_windows"]):
        if location_idx == data["depot"]:
            continue
        index = manager.NodeToIndex(location_idx)
        time_dimension.CumulVar(index).SetRange(time_window[0], time_window[1])
    # Add time window constraints for each vehicle start and end node.
    depot_idx = data["depot"]
    for vehicle_id in range(data["num_vehicles"]):
        index = routing.Start(vehicle_id)
        time_dimension.CumulVar(index).SetRange(
            data["time_windows"][depot_idx][0], data["time_windows"][depot_idx][1]
        )

    # Instantiate route start and end times to produce feasible times.
    for i in range(data["num_vehicles"]):
        routing.AddVariableMinimizedByFinalizer(
            time_dimension.CumulVar(routing.Start(i))
        )
        routing.AddVariableMinimizedByFinalizer(time_dimension.CumulVar(routing.End(i)))

    # Add Capacity constraint.
    def demand_callback(from_index):
        """Returns the demand of the node."""
        # Convert from routing variable Index to demands NodeIndex.
        from_node = manager.IndexToNode(from_index)
        return data["demands"][from_node]
    
    demand_callback_index = routing.RegisterUnaryTransitCallback(demand_callback)
    routing.AddDimensionWithVehicleCapacity(
        demand_callback_index,
        0, # slack
        data["vehicle_capacities"],  # vehicle maximum capacities
        True,  # start cumul to zero
        "Capacity",
    )

    # Define Transportation Requests.
    for index in range(0, len(data["pickups_deliveries"]), 2):
        request1 = data["pickups_deliveries"][index]
        pickup_index1 = manager.NodeToIndex(request1[0])
        delivery_index1 = manager.NodeToIndex(request1[1])
        routing.AddPickupAndDelivery(pickup_index1, delivery_index1)
        routing.solver().Add(
            routing.VehicleVar(pickup_index1) == routing.VehicleVar(delivery_index1)
        )
        routing.solver().Add(
            time_dimension.CumulVar(pickup_index1)
            <= time_dimension.CumulVar(delivery_index1)
        )

        request2 = data["pickups_deliveries"][index+1]
        pickup_index2 = manager.NodeToIndex(request2[0])
        delivery_index2 = manager.NodeToIndex(request2[1])
        routing.AddPickupAndDelivery(pickup_index2, delivery_index2)
        routing.solver().Add(
            routing.VehicleVar(pickup_index2) == routing.VehicleVar(delivery_index2)
        )
        routing.solver().Add(
            time_dimension.CumulVar(pickup_index2)
            <= time_dimension.CumulVar(delivery_index2)
        )

        # Make the same driver send a worker to work and back
        if (same_driver):
            routing.solver().Add(
                routing.VehicleVar(pickup_index1) == routing.VehicleVar(pickup_index2)
        )

    # Setting first solution heuristic.
    search_parameters = pywrapcp.DefaultRoutingSearchParameters()
    search_parameters.first_solution_strategy = (
        routing_enums_pb2.FirstSolutionStrategy.PARALLEL_CHEAPEST_INSERTION
    )

    # Solve the problem.
    solution = routing.SolveWithParameters(search_parameters)

    # Print solution on console.
    if solution:
        print_solution(data, manager, routing, solution, timezone)
        create_schedule(data, manager, routing, solution, output_file, timezone)
    else:
        print("\nNo solution\n")

    return solution


def create_schedule(data, manager, routing, solution, output_file, timezone):
    """Store schedule into Excel sheet."""
    wb = Workbook()
    ws = wb.active
    row_count = 1
    time_dimension = routing.GetDimensionOrDie("Time")
    heading_fill = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')
    subheading_fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
    for vehicle_id in range(data["num_vehicles"]):
        ws.append([data['driver_names'][vehicle_id]])
        print([data['driver_names'][vehicle_id]])
        ws.merge_cells(f'A{row_count}:E{row_count}')
        ws[f'A{row_count}'].fill = heading_fill
        row_count += 1
        ws.append(['Name', 'From', 'To',  'Pick-up Time', 'Arrival Time'])
        ws[f'A{row_count}'].fill = subheading_fill
        ws[f'B{row_count}'].fill = subheading_fill
        ws[f'C{row_count}'].fill = subheading_fill
        ws[f'D{row_count}'].fill = subheading_fill
        ws[f'E{row_count}'].fill = subheading_fill
        row_count += 1
        index = routing.Start(vehicle_id)
        while not routing.IsEnd(index):
            index = solution.Value(routing.NextVar(index))
            time_var = time_dimension.CumulVar(index)
            ind = manager.IndexToNode(index)
            row = []
            if data['node_type'][ind] == 'from_1' or data['node_type'][ind] == 'from_2':
                name = data['worker_name'][ind]
                home = data['addresses'][data['address_rpt'][ind][1]]
                if home in data['address_names']:
                    home = data['address_names'][home]
                dest = data['addresses'][data['address_rpt'][ind + 1][1]]
                if dest in data['address_names']:
                    dest = data['address_names'][dest]                
                home_address = urllib.parse.unquote(home)
                if home_address.endswith(', Kuwait'):
                    home_address = home_address[:-len(', Kuwait')]
                dest_address = urllib.parse.unquote(dest)
                if dest_address.endswith(', Kuwait'):
                    dest_address = dest_address[:-len(', Kuwait')]
                if data['node_type'][ind] == 'from_1':
                    # Pickup time range:
                    # pickup_time = f'{convert_seconds_to_hhmm(solution.Min(time_var))} to {convert_seconds_to_hhmm(solution.Max(time_var))}'
                    # Midpoint of pickup time range:
                    # pickup_time = convert_seconds_to_hhmm((solution.Min(time_var)+solution.Max(time_var)) // 2)
                    # Max pickup time:
                    pickup_time = epoch_time_to_time(solution.Max(time_var), timezone)
                    arrival_time = epoch_time_to_time(data['time_windows'][ind + 1][1], timezone)
                elif data['node_type'][ind] == 'from_2':
                    pickup_time = epoch_time_to_time((solution.Min(time_var)+solution.Max(time_var)) // 2, timezone)
                    arrival_time = '-'
                row = [ name,
                        home_address, 
                        dest_address, 
                        pickup_time, 
                        arrival_time]
                ws.append(row)
                print(row, '[', data['address_rpt'][ind][1], 'to', data['address_rpt'][ind + 1][1], ']')
                row_count += 1
        ws.append([])
        row_count += 1

    # Adjust column width
    for column_cells in ws.columns:
        # Get the maximum length of the content in the column
        max_length = 0
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Adjust the column width (adding a little extra space)
        adjusted_width = (max_length + 1)
        ws.column_dimensions[column_cells[1].column_letter].width = adjusted_width

    wb.save(output_file)

def main():
    """Entry point of the program."""

    # Read the Excel files
    drivers_df = pd.read_excel('Test_Drivers.xlsx')
    # drivers_df = pd.read_excel('Drivers.xlsx')
    workers_df = pd.read_excel('Test_Workers.xlsx')
    # workers_df = pd.read_excel('Workers_Test.xlsx')
    output_file = 'Schedule1.xlsx'
    vehicle_capacity = 3
    timezone = 'Asia/Kuwait'
    data = create_data_model(drivers_df, workers_df, vehicle_capacity, timezone)
    print(data)
    print_2d_matrix(data['time_matrix'])
    solve(data, same_driver=False, output_file=output_file)

if __name__ == "__main__":
    main()